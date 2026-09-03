"""
Sinh file Excel để người điền NỐT sáu trường tư vấn còn thiếu.

    python -m scripts.sinh_mau_bo_sung           ghi ra data/bo_sung_tu_van.xlsx
    python -m scripts.sinh_mau_bo_sung --ra <đường dẫn>

VÌ SAO CẦN SCRIPT NÀY
---------------------
Danh mục nạp từ Excel của cửa hàng chỉ có 9 trong 15 trường. Sáu trường
thiếu chính là phần TƯ VẤN — thứ ERP không giữ được và cũng là thứ làm nên
chất lượng câu trả lời:

    so_cong_bo    số công bố mỹ phẩm — PHÁP LÝ, khách và thanh tra đều hỏi
    khong_chua    khách dị ứng paraben/cồn hỏi "có chứa X không"
    hsd_thang     hạn dùng sau khi mở nắp
    cach_dung     dùng thế nào
    thoi_diem     sáng hay tối
    do_pH         hợp da nào

Thiếu chúng thì agent KHÔNG nói sai — nó chuyển cho người. Nhưng nó chuyển
cho người ở đúng những câu lẽ ra tự trả lời được, mỗi ngày, mãi mãi.

Bảo "thêm sáu cột vào Excel" là giao một việc mơ hồ: cột tên gì, viết dạng
gì, giá trị nào hợp lệ. File này sinh sẵn 13 dòng theo đúng mã hiện có, tiêu
đề đúng tên trường, kèm một dòng ví dụ và ghi chú ngay trong ô — người điền
chỉ việc gõ vào, không phải đoán.

KHÔNG TỰ ĐIỀN HỘ, VÀ ĐÓ LÀ CHỦ Ý
--------------------------------
Sáu trường này không suy ra được từ tên hàng. Số công bố mỹ phẩm là một con
số do Bộ Y tế cấp; đoán nó ra là bịa một dữ kiện pháp lý. `do_pH` đoán sai
thì agent tư vấn sai cho da nhạy cảm.

Nên script để trống, và `scripts.nap_catalog_tu_excel` cũng không bao giờ
tự sinh giá trị cho chúng.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

CATALOG = ROOT / "data" / "catalog.json"
CATALOG_MAU = ROOT / "data" / "catalog.example.json"
RA_MAC_DINH = ROOT / "data" / "bo_sung_tu_van.xlsx"

# Cột khoá — để đối chiếu, KHÔNG được sửa.
COT_KHOA = ("ma", "ten")

# Sáu trường cần điền, kèm hướng dẫn và ví dụ thật.
CAN_DIEN: tuple[tuple[str, str, str], ...] = (
    ("so_cong_bo",
     "Số công bố mỹ phẩm do Bộ Y tế cấp. BẮT BUỘC — khách và thanh tra đều hỏi. "
     "Chép nguyên văn trên phiếu công bố, không rút gọn.",
     "123456/22/CBMP-HN"),
    ("khong_chua",
     "Những thứ sản phẩm KHÔNG chứa, cách nhau bằng dấu phẩy. Dùng để trả lời "
     "khách dị ứng. Không chắc thì để trống, đừng đoán.",
     "paraben, cồn khô, hương liệu"),
    ("hsd_thang",
     "Hạn dùng sau khi MỞ NẮP, tính bằng tháng. Chỉ ghi số.",
     "12"),
    ("cach_dung",
     "Một tới hai câu, đúng thứ tự các bước.",
     "Lấy lượng vừa đủ, thoa đều lên da khô, massage 30 giây rồi rửa lại."),
    ("thoi_diem",
     "Dùng lúc nào: sáng, tối, hoặc cả hai.",
     "tối"),
    ("do_pH",
     "Độ pH của sản phẩm. Chỉ ghi số, dùng dấu chấm thập phân.",
     "5.5"),
)

VANG = PatternFill("solid", fgColor="FFF2CC")
XAM = PatternFill("solid", fgColor="E7E6E6")
XANH = PatternFill("solid", fgColor="DDEBF7")


def _doc_danh_muc() -> list[dict]:
    """Danh mục thật, rơi về bản mẫu như mọi chỗ khác đọc tệp này."""
    tep = CATALOG if CATALOG.exists() else CATALOG_MAU
    if not tep.exists():
        raise SystemExit(
            "Không có data/catalog.json lẫn catalog.example.json. "
            "Chạy `python -m scripts.san_sang` để biết còn thiếu gì."
        )
    return json.loads(tep.read_text(encoding="utf-8")).get("san_pham", [])


def dung_workbook(san_pham: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bổ sung tư vấn"

    tieu_de = [*COT_KHOA, *(t for t, _, _ in CAN_DIEN)]

    # Dòng 1: tên trường — đây là thứ `nap_catalog_tu_excel` đọc THEO TÊN,
    # nên tuyệt đối không đổi. Dòng 2: hướng dẫn cho người.
    ws.append(tieu_de)
    ws.append(["(không sửa)", "(không sửa)", *(h for _, h, _ in CAN_DIEN)])
    ws.append(["VÍ DỤ", "Serum mẫu", *(v for _, _, v in CAN_DIEN)])

    for i, _ in enumerate(tieu_de, start=1):
        o = ws.cell(row=1, column=i)
        o.font = Font(bold=True)
        o.fill = XAM if i <= len(COT_KHOA) else VANG
        o.alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=i).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=2, column=i).font = Font(size=9, italic=True, color="666666")
        ws.cell(row=3, column=i).fill = XANH
        ws.cell(row=3, column=i).font = Font(size=9, italic=True)

    for sp in san_pham:
        ws.append([sp.get("ma", ""), sp.get("ten", ""), *[""] * len(CAN_DIEN)])

    # Khoá hai cột đầu để không ai sửa nhầm mã — nạp lại theo mã, sửa mã là
    # dòng đó thành sản phẩm mới và bản ghi cũ mất hồ sơ tư vấn.
    ws.freeze_panes = "C4"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 46
    for i in range(len(COT_KHOA) + 1, len(tieu_de) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 30
    ws.row_dimensions[2].height = 74

    return wb


def main() -> None:
    ra = RA_MAC_DINH
    if "--ra" in sys.argv:
        ra = Path(sys.argv[sys.argv.index("--ra") + 1])

    san_pham = _doc_danh_muc()
    if not san_pham:
        raise SystemExit("Danh mục rỗng — không có mã nào để dựng mẫu.")

    ra.parent.mkdir(parents=True, exist_ok=True)
    dung_workbook(san_pham).save(ra)

    print(f"Đã ghi {ra}")
    print(f"  {len(san_pham)} sản phẩm · {len(CAN_DIEN)} cột cần điền\n")
    print("Điền xong thì nạp lại bằng:")
    print("  python -m scripts.nap_catalog_tu_excel <đường dẫn> --ghi\n")
    print("Dòng 2 là hướng dẫn, dòng 3 là ví dụ — XOÁ CẢ HAI trước khi nạp.")
    print("Ô nào chưa chắc thì để TRỐNG. Trống thì agent chuyển cho người;")
    print("điền bừa thì agent nói sai một cách rất tự tin.")


if __name__ == "__main__":
    main()
