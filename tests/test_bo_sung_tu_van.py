"""
Gộp sáu trường tư vấn từ tệp Excel bổ sung.

BA KHẲNG ĐỊNH QUAN TRỌNG NHẤT, cả ba đều canh một kiểu hỏng IM LẶNG:

  1. Dòng VÍ DỤ trong mẫu không được nạp. Người dùng quên xoá nó thì mọi
     sản phẩm mang số công bố "123456/22/CBMP-HN" — một dữ kiện PHÁP LÝ
     sai, gán cho cả danh mục, và agent đọc ra rất tự tin.

  2. Ô TRỐNG không ghi đè giá trị đang có. Ô trống nghĩa là "chưa điền",
     không phải "xoá đi". Không có chốt này thì mở tệp sửa một dòng rồi nạp
     lại sẽ xoá sạch mười hai dòng kia.

  3. Mã lạ không âm thầm biến mất. Gõ sai mã thì phải BÁO, chứ không phải
     nạp vào hư không rồi báo "xong".
"""
from __future__ import annotations

import pathlib
import sys

import pytest

ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from scripts.nap_bo_sung_tu_van import (  # noqa: E402
    TEN_TRUONG,
    LoiBoSung,
    doc,
    gop,
)
from scripts.sinh_mau_bo_sung import CAN_DIEN, dung_workbook  # noqa: E402


def _mau(tmp_path, san_pham=None):
    """Sinh mẫu thật rồi trả về (đường dẫn, chỉ số cột)."""
    sp = san_pham or [
        {"ma": "A-1", "ten": "Serum A"},
        {"ma": "B-2", "ten": "Kem B"},
    ]
    tep = tmp_path / "bo_sung.xlsx"
    dung_workbook(sp).save(tep)
    ws = load_workbook(tep).active
    cot = {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}
    return tep, cot


def _dien(tep, dong, cot, **gia_tri):
    wb = load_workbook(tep)
    ws = wb.active
    for ten, gt in gia_tri.items():
        ws.cell(dong, cot[ten], gt)
    wb.save(tep)


# ---------------------------------------------------------------
#  1. Dòng ví dụ và hướng dẫn không được nạp
# ---------------------------------------------------------------

def test_dong_vi_du_khong_bao_gio_duoc_nap(tmp_path):
    """
    Khẳng định quan trọng nhất tệp này. Mẫu sinh ra CÓ dòng ví dụ chứa một
    số công bố giả — nạp nó vào là gán một dữ kiện pháp lý sai.
    """
    tep, _ = _mau(tmp_path)
    ra = doc(tep)
    assert "VÍ DỤ" not in ra
    assert not any("123456" in str(v) for v in ra.values())


def test_mau_moi_sinh_ra_thi_khong_co_du_lieu_nao(tmp_path):
    """Chưa ai điền gì thì gộp phải là việc rỗng, không phải điền chuỗi rỗng."""
    tep, _ = _mau(tmp_path)
    assert doc(tep) == {}


def test_mau_co_du_sau_cot_can_dien(tmp_path):
    tep, cot = _mau(tmp_path)
    for ten in TEN_TRUONG:
        assert ten in cot, f"mẫu thiếu cột {ten}"
    assert "ma" in cot and "ten" in cot


# ---------------------------------------------------------------
#  2. Ô trống không ghi đè
# ---------------------------------------------------------------

def test_o_trong_khong_ghi_de_gia_tri_dang_co(tmp_path):
    """
    Ô trống nghĩa là "chưa điền", KHÔNG phải "xoá đi". Thiếu chốt này thì
    mở tệp sửa một dòng rồi nạp lại sẽ xoá sạch những dòng kia.
    """
    tep, cot = _mau(tmp_path)
    _dien(tep, 4, cot, so_cong_bo="111/24/CBMP-HN")

    danh_muc = [
        {"ma": "A-1", "ten": "Serum A", "do_pH": 5.5, "hsd_thang": 12},
        {"ma": "B-2", "ten": "Kem B", "so_cong_bo": "CU-999"},
    ]
    dien, bo_qua, la = gop(danh_muc, doc(tep))
    assert danh_muc[0]["do_pH"] == 5.5, "ô trống đã xoá mất giá trị cũ"
    assert danh_muc[0]["hsd_thang"] == 12
    assert danh_muc[1]["so_cong_bo"] == "CU-999", "B-2 không điền mà bị đụng"
    assert danh_muc[0]["so_cong_bo"] == "111/24/CBMP-HN"
    assert dien == 1


def test_gia_tri_da_co_thi_giu_tru_khi_bao_ghi_de(tmp_path):
    tep, cot = _mau(tmp_path)
    _dien(tep, 4, cot, so_cong_bo="MOI-111")

    danh_muc = [{"ma": "A-1", "ten": "Serum A", "so_cong_bo": "CU-999"}]
    dien, bo_qua, _ = gop(danh_muc, doc(tep))
    assert danh_muc[0]["so_cong_bo"] == "CU-999"
    assert (dien, bo_qua) == (0, 1)

    gop(danh_muc, doc(tep), ghi_de=True)
    assert danh_muc[0]["so_cong_bo"] == "MOI-111"


# ---------------------------------------------------------------
#  3. Mã lạ phải được BÁO
# ---------------------------------------------------------------

def test_ma_khong_co_trong_danh_muc_thi_bao_chu_khong_im(tmp_path):
    """
    Gõ sai mã mà script báo "xong" là kiểu hỏng im lặng: người vận hành
    tưởng đã điền, agent thì vẫn không có gì.
    """
    tep, cot = _mau(tmp_path)
    _dien(tep, 4, cot, so_cong_bo="111")

    _, _, la = gop([{"ma": "KHAC", "ten": "X"}], doc(tep))
    assert la == ["A-1"]


# ---------------------------------------------------------------
#  Chuẩn hoá kiểu dữ liệu
# ---------------------------------------------------------------

def test_khong_chua_tach_thanh_danh_sach(tmp_path):
    tep, cot = _mau(tmp_path)
    _dien(tep, 4, cot, khong_chua="paraben, cồn khô , hương liệu")
    assert doc(tep)["A-1"]["khong_chua"] == ["paraben", "cồn khô", "hương liệu"]


def test_hsd_thang_thanh_so_nguyen(tmp_path):
    tep, cot = _mau(tmp_path)
    _dien(tep, 4, cot, hsd_thang="12")
    assert doc(tep)["A-1"]["hsd_thang"] == 12


def test_do_pH_nhan_ca_dau_phay_thap_phan(tmp_path):
    """Người Việt gõ 5,5 chứ không phải 5.5 — nhận cả hai."""
    tep, cot = _mau(tmp_path)
    _dien(tep, 4, cot, do_pH="5,5")
    assert doc(tep)["A-1"]["do_pH"] == 5.5


def test_so_sai_dinh_dang_thi_NEM_chu_khong_bo_qua(tmp_path):
    """
    Bỏ qua trong im lặng là người vận hành gõ "mười hai tháng" rồi tưởng đã
    điền xong.
    """
    tep, cot = _mau(tmp_path)
    _dien(tep, 4, cot, hsd_thang="mười hai")
    with pytest.raises(LoiBoSung, match="phải là số"):
        doc(tep)


def test_khoang_trang_khong_thanh_gia_tri(tmp_path):
    """Ô chỉ có dấu cách phải coi như trống, không phải chuỗi rỗng."""
    tep, cot = _mau(tmp_path)
    _dien(tep, 4, cot, so_cong_bo="   ")
    assert doc(tep) == {}


# ---------------------------------------------------------------
#  Hai script nói cùng một bộ trường
# ---------------------------------------------------------------

def test_hai_script_dung_chung_mot_danh_sach_truong():
    """
    Sinh mẫu và nạp mẫu lệch nhau một tên cột thì cột đó im lặng không bao
    giờ được nạp — người điền đầy đủ mà agent vẫn không biết gì.
    """
    assert TEN_TRUONG == tuple(t for t, _, _ in CAN_DIEN)


def test_sau_truong_dung_la_sau_truong_catalog_con_thieu():
    """
    Danh sách này phải khớp `THIEU` trong `nap_catalog_tu_excel` — đó là
    nơi khai rằng bảng hàng của cửa hàng không có chúng.
    """
    from scripts.nap_catalog_tu_excel import THIEU

    assert set(TEN_TRUONG) == set(THIEU)


def test_thieu_tep_thi_bao_cach_sinh(tmp_path):
    with pytest.raises(LoiBoSung, match="sinh_mau_bo_sung"):
        doc(tmp_path / "khong-co.xlsx")


# ---------------------------------------------------------------
#  Không lọt dữ liệu doanh nghiệp lên GitHub
# ---------------------------------------------------------------

def test_tep_bo_sung_bi_gitignore_chan():
    """
    Điền xong, tệp này chứa SỐ CÔNG BỐ MỸ PHẨM và thành phần thật — đúng
    loại dữ liệu `data/catalog.json` đang bị chặn, chỉ khác định dạng.

    Repo này đã một lần chặn danh mục mà quên bản sao lưu của chính nó, và
    bản sao ấy lọt lên GitHub. Chặn NGAY khi sinh ra tệp, không đợi tới lúc
    nó có dữ liệu.
    """
    ignore = (ROOT / ".gitignore").read_text(encoding="utf-8")
    assert "data/bo_sung_tu_van.xlsx" in ignore, (
        "Tệp bổ sung chưa bị .gitignore chặn — điền xong là dữ liệu doanh "
        "nghiệp lên GitHub"
    )


def test_duong_dan_mac_dinh_khop_voi_dong_trong_gitignore():
    """
    Đổi đường dẫn mặc định trong script mà quên sửa .gitignore là chặn một
    tệp không còn ai sinh ra, còn tệp thật thì lên repo.
    """
    from scripts.sinh_mau_bo_sung import RA_MAC_DINH

    tuong_doi = RA_MAC_DINH.relative_to(ROOT).as_posix()
    ignore = (ROOT / ".gitignore").read_text(encoding="utf-8")
    assert tuong_doi in ignore, (
        f"{tuong_doi} là đường dẫn script ghi ra, nhưng .gitignore không có dòng đó"
    )
