"""
Gộp sáu trường tư vấn từ `bo_sung_tu_van.xlsx` vào `data/catalog.json`.

    python -m scripts.nap_bo_sung_tu_van                 xem trước, KHÔNG ghi
    python -m scripts.nap_bo_sung_tu_van --ghi           ghi thật
    python -m scripts.nap_bo_sung_tu_van --tep <đường dẫn> --ghi

VÌ SAO LÀ SCRIPT RIÊNG, KHÔNG NHÉT VÀO `nap_catalog_tu_excel`
-------------------------------------------------------------
Hai tệp Excel có hai chủ sở hữu khác nhau.

`nap_catalog_tu_excel` đọc bảng hàng của CỬA HÀNG: mã, tên, giá, tồn — thứ
xuất ra từ phần mềm bán hàng và sẽ được xuất lại mỗi lần đổi bảng giá. Nó cố
ý liệt sáu trường tư vấn vào `THIEU` và để trống, vì bảng ấy không có chúng
và đoán ra là bịa.

Sáu trường tư vấn thì do người của cửa hàng ngồi viết một lần, và không đổi
theo bảng giá. Bắt chúng đi chung một tệp nghĩa là mỗi lần xuất lại bảng giá
là mất sạch phần tư vấn — hoặc phải nhớ chép tay sang tệp mới, và sẽ có ngày
quên.

Tách hai đường nạp thì xuất lại bảng giá bao nhiêu lần cũng được, phần tư vấn
nằm yên.

KHÔNG BAO GIỜ GHI ĐÈ GIÁ TRỊ ĐANG CÓ, TRỪ KHI ĐƯỢC BẢO
------------------------------------------------------
Ô trống trong tệp bổ sung nghĩa là "chưa điền", KHÔNG phải "xoá đi". Gộp mà
để ô trống ghi đè lên giá trị cũ thì mở tệp lên sửa một dòng rồi nạp lại sẽ
xoá sạch mười hai dòng kia — im lặng.
"""
from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

GOC = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(GOC))

from openpyxl import load_workbook  # noqa: E402

from scripts.sinh_mau_bo_sung import CAN_DIEN  # noqa: E402

CATALOG = GOC / "data" / "catalog.json"
TEP_MAC_DINH = GOC / "data" / "bo_sung_tu_van.xlsx"

TEN_TRUONG = tuple(t for t, _, _ in CAN_DIEN)

# Trường là DANH SÁCH: trong Excel viết liền, cách nhau bởi dấu phẩy.
DANH_SACH = {"khong_chua"}
# Trường là SỐ.
SO_NGUYEN = {"hsd_thang"}
SO_THUC = {"do_pH"}

# Dòng ví dụ và dòng hướng dẫn do `sinh_mau_bo_sung` chèn vào. Nếu người
# dùng quên xoá, nạp chúng vào là mọi sản phẩm mang số công bố "123456/22/
# CBMP-HN" — một dữ kiện PHÁP LÝ sai, gán cho cả danh mục, rất tự tin.
DONG_BO_QUA = {"(không sửa)", "ví dụ", "vi du"}


class LoiBoSung(ValueError):
    """Tệp bổ sung sai. Thông điệp nói rõ sửa gì."""


def _chuan(ten: str, tho, ma: str):
    """Đổi một ô Excel thành giá trị đúng kiểu. Ô trống -> None (bỏ qua)."""
    if tho is None:
        return None
    s = str(tho).strip()
    if not s:
        return None

    if ten in SO_NGUYEN:
        try:
            return int(float(s))
        except ValueError as exc:
            raise LoiBoSung(
                f"{ma}: cột {ten!r} phải là số, đang là {s!r}."
            ) from exc
    if ten in SO_THUC:
        try:
            return round(float(s.replace(",", ".")), 2)
        except ValueError as exc:
            raise LoiBoSung(
                f"{ma}: cột {ten!r} phải là số, đang là {s!r}."
            ) from exc
    if ten in DANH_SACH:
        return [x.strip() for x in s.split(",") if x.strip()]
    return s


def doc(tep: Path) -> dict[str, dict]:
    """Đọc tệp bổ sung thành {mã: {trường: giá trị}}. Chỉ ô ĐÃ ĐIỀN."""
    if not tep.exists():
        raise LoiBoSung(
            f"Không có {tep}. Sinh mẫu bằng:\n"
            "  python -m scripts.sinh_mau_bo_sung"
        )
    ws = load_workbook(tep, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise LoiBoSung(f"{tep.name} rỗng.")

    tieu_de = [str(x).strip() if x else "" for x in rows[0]]
    if "ma" not in tieu_de:
        raise LoiBoSung(
            f"{tep.name} thiếu cột 'ma' ở dòng đầu. Đừng sửa dòng tiêu đề — "
            "sinh lại mẫu bằng `python -m scripts.sinh_mau_bo_sung`."
        )
    i_ma = tieu_de.index("ma")
    co = {t: tieu_de.index(t) for t in TEN_TRUONG if t in tieu_de}
    if not co:
        raise LoiBoSung(
            f"{tep.name} không có cột nào trong sáu cột cần điền: "
            + ", ".join(TEN_TRUONG)
        )

    ra: dict[str, dict] = {}
    for r in rows[1:]:
        if not r or i_ma >= len(r) or not r[i_ma]:
            continue
        ma = str(r[i_ma]).strip()
        if not ma or ma.lower() in DONG_BO_QUA:
            continue
        ban_ghi = {}
        for ten, i in co.items():
            if i >= len(r):
                continue
            gt = _chuan(ten, r[i], ma)
            if gt is not None and gt != []:
                ban_ghi[ten] = gt
        if ban_ghi:
            ra[ma] = ban_ghi
    return ra


def gop(danh_muc: list[dict], bo_sung: dict[str, dict],
        ghi_de: bool = False) -> tuple[int, int, list[str]]:
    """Gộp tại chỗ. Trả (số ô điền, số ô bỏ qua, mã lạ không có trong danh mục)."""
    theo_ma = {sp.get("ma"): sp for sp in danh_muc}
    la = [m for m in bo_sung if m not in theo_ma]
    dien = bo_qua = 0
    for ma, ban_ghi in bo_sung.items():
        sp = theo_ma.get(ma)
        if sp is None:
            continue
        for ten, gt in ban_ghi.items():
            cu = sp.get(ten)
            # Có giá trị rồi thì GIỮ, trừ khi được bảo ghi đè. Xem đầu tệp.
            if cu not in (None, "", [], 0) and not ghi_de:
                bo_qua += 1
                continue
            sp[ten] = gt
            dien += 1
    return dien, bo_qua, la


def main(argv: list[str] | None = None) -> int:
    argv = sys.argv[1:] if argv is None else argv
    tep = TEP_MAC_DINH
    if "--tep" in argv:
        tep = Path(argv[argv.index("--tep") + 1])
    ghi = "--ghi" in argv
    ghi_de = "--ghi-de" in argv

    if not CATALOG.exists():
        print(f"Không có {CATALOG}. Nạp danh mục trước.")
        return 1

    try:
        bo_sung = doc(tep)
    except LoiBoSung as exc:
        print(f"LỖI: {exc}")
        return 1

    goc = json.loads(CATALOG.read_text(encoding="utf-8"))
    danh_muc = goc.get("san_pham", [])
    dien, bo_qua, la = gop(danh_muc, bo_sung, ghi_de)

    print(f"Tệp bổ sung : {tep.name} — {len(bo_sung)} mã có dữ liệu")
    print(f"Sẽ điền     : {dien} ô")
    if bo_qua:
        print(f"Giữ nguyên  : {bo_qua} ô đã có giá trị (dùng --ghi-de để đè)")
    if la:
        print(f"Mã KHÔNG có trong danh mục ({len(la)}): " + ", ".join(la[:8]))

    # Còn thiếu bao nhiêu — con số này là thứ đáng nhìn nhất.
    con_trong: dict[str, int] = {t: 0 for t in TEN_TRUONG}
    for sp in danh_muc:
        for t in TEN_TRUONG:
            if sp.get(t) in (None, "", []):
                con_trong[t] += 1
    print("\nSau khi gộp, còn trống:")
    for t, n in con_trong.items():
        dau = "  ✓" if n == 0 else "   "
        print(f"{dau} {t:14} {n}/{len(danh_muc)}")

    if not ghi:
        print("\nXem trước, CHƯA ghi. Thêm --ghi để ghi thật.")
        return 0

    sao_luu = CATALOG.with_suffix(".json.bak")
    shutil.copy2(CATALOG, sao_luu)
    goc["san_pham"] = danh_muc
    CATALOG.write_text(
        json.dumps(goc, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"\nĐã ghi {CATALOG} (bản cũ: {sao_luu.name})")
    print("Chạy `python -m scripts.san_sang` để kiểm lại.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
