"""
Nạp danh mục sản phẩm từ tệp Excel vào `data/catalog.json`.

    python -m scripts.nap_catalog_tu_excel <tệp.xlsx>            xem trước
    python -m scripts.nap_catalog_tu_excel <tệp.xlsx> --ghi      ghi thật

VÌ SAO ĐỌC CỘT THEO TÊN TIÊU ĐỀ, KHÔNG THEO VỊ TRÍ
--------------------------------------------------
Bảng giá là tệp NGƯỜI sửa. Ai đó chèn thêm một cột giữa chừng là mọi vị trí
lệch một ô — và lệch im lặng: giá đọc thành tồn kho, tồn kho đọc thành mô
tả. Không có gì nổ, chỉ có một danh mục sai toàn bộ.

Đọc theo tên thì chèn cột không ảnh hưởng, và đổi tên cột thì báo lỗi ngay.

VÌ SAO KHÔNG TỰ ĐIỀN TRƯỜNG THIẾU
---------------------------------
Bảng Excel không có pH, số công bố, hạn dùng, cách dùng. Đoán chúng là bịa
ra căn cứ cho agent trích dẫn. Script bỏ trống và NÓI RA còn thiếu gì.
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
from pathlib import Path

# Console Windows mac dinh cp1258 khong in duoc tieng Viet.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

GOC = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(GOC))

DICH = GOC / "data" / "catalog.json"

# Tên cột trong tệp -> khoá trong catalog. Thiếu cột nào thì báo, không đoán.
COT = {
    "Mã sản phẩm (SKU)": "ma",
    "Tên sản phẩm": "ten",
    "Phân loại": "loai",
    "Dung tích / Quy cách": "dung_tich",
    "Đơn giá niêm yết (VNĐ)": "gia",
    "Tồn kho khả dụng": "ton_kho",
    "Loại da phù hợp": "da_phu_hop",
    "Vấn đề hỗ trợ / Công dụng": "van_de_ho_tro",
    "Thành phần chính": "thanh_phan_chinh",
}
# Ba trường này là DANH SÁCH, trong Excel viết liền nhau cách bởi dấu phẩy.
DANH_SACH = {"da_phu_hop", "van_de_ho_tro", "thanh_phan_chinh"}
# Có trong lược đồ nhưng KHÔNG có trong bảng Excel. Bỏ trống, không đoán.
THIEU = ("khong_chua", "do_pH", "cach_dung", "thoi_diem",
         "so_cong_bo", "hsd_thang")


def _tach(gia_tri) -> list[str]:
    return [p.strip() for p in str(gia_tri or "").split(",") if p.strip()]


def _so(gia_tri, ten_cot: str, ma: str) -> int:
    try:
        return int(float(str(gia_tri).replace(",", "").replace(".", "").strip()
                         or 0))
    except ValueError:
        raise SystemExit(
            f"[LỖI] {ma}: cột {ten_cot!r} không phải số: {gia_tri!r}"
        ) from None


def doc(tep: Path) -> tuple[list[dict], str]:
    import openpyxl

    wb = openpyxl.load_workbook(tep, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Tìm dòng tiêu đề thay vì gắn cứng "dòng 3": tệp người sửa hay có thêm
    # dòng tiêu đề lớn, ghi chú, dòng trống ở đầu.
    i_dau = None
    for i, r in enumerate(rows):
        if r and "Mã sản phẩm (SKU)" in [str(x).strip() for x in r if x]:
            i_dau = i
            break
    if i_dau is None:
        raise SystemExit("[LỖI] Không thấy dòng tiêu đề có cột 'Mã sản phẩm (SKU)'")

    tieu_de = [str(x).strip() if x else "" for x in rows[i_dau]]
    thieu_cot = [c for c in COT if c not in tieu_de]
    if thieu_cot:
        raise SystemExit("[LỖI] Tệp thiếu cột: " + " · ".join(thieu_cot))
    vi_tri = {tieu_de.index(c): k for c, k in COT.items()}

    san_pham = []
    for r in rows[i_dau + 1:]:
        if not r or not r[tieu_de.index("Mã sản phẩm (SKU)")]:
            continue
        sp: dict = {}
        for i, khoa in vi_tri.items():
            gia_tri = r[i]
            if khoa in DANH_SACH:
                sp[khoa] = _tach(gia_tri)
            elif khoa in ("gia", "ton_kho"):
                sp[khoa] = _so(gia_tri, khoa, str(r[1]))
            else:
                sp[khoa] = str(gia_tri or "").strip()
        san_pham.append(sp)

    # Thương hiệu lấy từ dòng tiêu đề lớn nếu có, để không phải gõ tay.
    thuong_hieu = ""
    for r in rows[:i_dau]:
        for x in (r or ()):
            if x and "BLANICA" in str(x).upper():
                thuong_hieu = "BLANICA"
    return san_pham, thuong_hieu


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Nạp danh mục từ Excel")
    p.add_argument("tep")
    p.add_argument("--ghi", action="store_true",
                   help="ghi thật vào data/catalog.json (mặc định chỉ xem)")
    a = p.parse_args(argv)

    tep = Path(a.tep)
    if not tep.exists():
        print(f"[LỖI] Không thấy {tep}")
        return 1

    san_pham, thuong_hieu = doc(tep)
    if not san_pham:
        print("[LỖI] Không đọc được sản phẩm nào.")
        return 1

    ma = [sp["ma"] for sp in san_pham]
    if len(ma) != len(set(ma)):
        trung = sorted({m for m in ma if ma.count(m) > 1})
        print(f"[LỖI] Mã trùng trong tệp: {trung}")
        return 1

    print(f"Đọc {len(san_pham)} sản phẩm · thương hiệu {thuong_hieu or '(chưa rõ)'}\n")
    for sp in san_pham:
        print(f"  {sp['ma']:<26} {sp['gia']:>9,}đ · tồn {sp['ton_kho']:<4}"
              f" · {len(sp['da_phu_hop'])} loại da"
              f" · {len(sp['thanh_phan_chinh'])} thành phần")

    print("\nTrường KHÔNG có trong tệp, để trống chứ không đoán:")
    print("  " + " · ".join(THIEU))

    if not a.ghi:
        print("\nChế độ xem — chưa ghi gì. Thêm --ghi để ghi thật.")
        return 0

    if DICH.exists():
        sao_luu = DICH.with_suffix(".json.bak")
        shutil.copyfile(DICH, sao_luu)
        print(f"\nĐã sao lưu danh mục cũ: {sao_luu.name}")

    DICH.write_text(
        json.dumps(
            {
                "_ghi_chu": (
                    f"Nạp từ {tep.name} bằng scripts.nap_catalog_tu_excel. "
                    "Sửa bảng Excel rồi chạy lại thay vì sửa tay tệp này."
                ),
                # KHÔNG đặt cờ `du_lieu_mau`: đây là hàng thật. Cờ đó là dấu
                # hiệu do người tạo dữ liệu tự khai, và khai sai là để
                # `scripts/san_sang.py` cho qua một danh mục bịa.
                "thuong_hieu": thuong_hieu or "BLANICA",
                "san_pham": san_pham,
                "don_hang": [],
            },
            ensure_ascii=False, indent=2,
        ) + "\n",
        encoding="utf-8",
    )
    print(f"Đã ghi {DICH.relative_to(GOC)}: {len(san_pham)} sản phẩm.")
    print(
        "\nBƯỚC TIẾP THEO:\n"
        "  python -m scripts.san_sang                    kiểm còn việc CHẶN nào\n"
        "  python -m scripts.nap_san_pham_erp --nhap-ton đẩy lên ERPNext"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
